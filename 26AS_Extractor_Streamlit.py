import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="26AS Extractor", layout="wide")

# ------------------------------------------------------------------
# Background - professional slate-blue theme
# ------------------------------------------------------------------
st.markdown(
    """
    <style>

    /* 1. PAGE CANVAS */
    [data-testid="stAppViewContainer"],
    .stApp {
        background: linear-gradient(135deg, #0f1c2e 0%, #142540 40%, #0a1628 100%);
        background-attachment: fixed;
    }

    /* 2. HEADER BAR */
    [data-testid="stHeader"] {
        background: rgba(10, 22, 40, 0.92) !important;
        border-bottom: 1px solid rgba(99, 179, 237, 0.15);
        backdrop-filter: blur(10px);
    }

    /* 3. SIDEBAR */
    [data-testid="stSidebar"] {
        background: #0d1f35 !important;
        border-right: 1px solid rgba(99, 179, 237, 0.12);
    }

    /* 4. MAIN CONTENT CARD */
    [data-testid="block-container"] {
        background: rgba(15, 30, 52, 0.82);
        border: 1px solid rgba(99, 179, 237, 0.18);
        border-radius: 16px;
        padding: 2.2rem 2.8rem;
        backdrop-filter: blur(12px);
        box-shadow: 0 8px 40px rgba(0, 0, 0, 0.55);
        margin-top: 1rem;
    }

    /* 5. ALL TEXT */
    html, body,
    .stApp, .stMarkdown,
    [class*="css"], p, span, div, label {
        color: #e2eaf4 !important;
    }

    /* 6. TITLE */
    h1 {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 2rem !important;
        letter-spacing: -0.5px;
        padding-bottom: 0.25rem;
        border-bottom: 2px solid #2b6cb0;
        margin-bottom: 1.5rem !important;
    }

    /* 7. SUBHEADERS */
    h2, h3 {
        color: #90cdf4 !important;
        font-weight: 600 !important;
        margin-top: 1.6rem !important;
    }

    /* 8. FILE UPLOADER - label */
    [data-testid="stFileUploader"] {
        background: rgba(43, 108, 176, 0.12) !important;
        border: 2px dashed #2b6cb0 !important;
        border-radius: 12px !important;
        padding: 1rem !important;
    }
    [data-testid="stFileUploader"] label p,
    [data-testid="stFileUploaderDropzoneInstructions"] span,
    [data-testid="stFileUploaderDropzoneInstructions"] p,
    [data-testid="stFileUploader"] section > div > span,
    [data-testid="stFileUploader"] section label {
        color: #90cdf4 !important;
        font-size: 0.95rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.3px;
    }
    [data-testid="stFileUploader"] button {
        background-color: #2b6cb0 !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
    }
    [data-testid="stFileUploader"] button:hover {
        background-color: #3182ce !important;
    }
    [data-testid="stFileUploader"] small,
    [data-testid="stFileUploaderDropzoneInstructions"] small {
        color: #a0c4e4 !important;
        font-size: 0.82rem !important;
    }

    /* 9. DATAFRAME */
    [data-testid="stDataFrame"] {
        border: 1px solid rgba(99, 179, 237, 0.2) !important;
        border-radius: 10px !important;
        overflow: hidden;
    }

    /* 10. DOWNLOAD BUTTON */
    [data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #1f4e78, #2b6cb0) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.6rem 1.8rem !important;
        font-size: 1rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.4px;
        box-shadow: 0 4px 14px rgba(43, 108, 176, 0.45);
        transition: all 0.2s ease;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: linear-gradient(135deg, #2b6cb0, #3182ce) !important;
        box-shadow: 0 6px 20px rgba(49, 130, 206, 0.55);
        transform: translateY(-1px);
    }

    /* 11. SPINNER */
    [data-testid="stSpinner"] p {
        color: #90cdf4 !important;
        font-weight: 500;
    }

    /* 12. ALERTS */
    [data-testid="stAlert"] {
        background: rgba(43, 108, 176, 0.15) !important;
        border-left: 4px solid #3182ce !important;
        border-radius: 8px !important;
        color: #e2eaf4 !important;
    }

    </style>
    """,
    unsafe_allow_html=True,
)

st.title("26AS PDF → Excel Extractor")

uploaded_file = st.file_uploader(
    "Upload Form 26AS PDF",
    type=["pdf"]
)

# ==============================================================
# HELPER FUNCTIONS
# ==============================================================

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT  = Font(bold=True)
NUM_FORMAT  = '#,##0.00'

_table_counter = [0]          # mutable so nested helpers can bump it


def _next_table_name(prefix: str) -> str:
    """Return a unique Excel Table display-name."""
    _table_counter[0] += 1
    safe = re.sub(r'[^A-Za-z0-9]', '_', prefix)
    return f"{safe}_{_table_counter[0]}"


def apply_sheet_formatting(
    ws,
    numeric_cols: list = None,
    total_col_indices: list = None,
    table_name: str = None,
):
    """
    Apply standard formatting to a worksheet:
      • Blue header row with white bold text
      • Auto column widths
      • Freeze pane at A2
      • Optional total row (bold) at bottom
      • Optional number formatting on specified 1-based column indices
      • Optional Excel Table
    """
    ws.freeze_panes = "A2"

    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=False)

    max_data_row = ws.max_row

    # Total row
    if total_col_indices:
        total_row = max_data_row + 1
        ws.cell(total_row, 1).value = "TOTAL"
        for col_idx, col_letter in total_col_indices:
            ws.cell(total_row, col_idx).value = (
                f"=SUM({col_letter}2:{col_letter}{max_data_row})"
            )
        for c in ws[total_row]:
            c.font = TOTAL_FONT
        max_data_row = total_row

    # Number formatting
    if numeric_cols:
        for col_idx in numeric_cols:
            for row in ws.iter_rows(min_row=2, max_row=max_data_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = NUM_FORMAT

    # Auto column widths
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 3
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 60)

    # Excel Table
    if table_name and max_data_row > 1:
        last_col_letter = get_column_letter(ws.max_column)
        table_ref = f"A1:{last_col_letter}{max_data_row}"
        tbl = Table(displayName=table_name, ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)


def write_sheet(writer, df: pd.DataFrame, sheet_name: str,
                numeric_col_names: list = None,
                add_totals: bool = False,
                table_prefix: str = None):
    """
    Write a DataFrame to an Excel sheet and apply standard formatting.

    numeric_col_names  – column names that should receive number formatting
    add_totals         – add a TOTAL row for all numeric-formatted columns
    table_prefix       – if supplied, an Excel Table is created
    """
    if df is None or df.empty:
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # Build 1-based column-index lists from column names
    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}

    num_indices = []
    if numeric_col_names:
        num_indices = [col_map[n] for n in numeric_col_names if n in col_map]

    total_pairs = []
    if add_totals and num_indices:
        total_pairs = [(i, get_column_letter(i)) for i in num_indices]

    tname = _next_table_name(table_prefix) if table_prefix else None

    apply_sheet_formatting(
        ws,
        numeric_cols=num_indices if num_indices else None,
        total_col_indices=total_pairs if total_pairs else None,
        table_name=tname,
    )


def safe_float(val: str) -> float:
    """Convert a string with commas to float, return 0.0 on failure."""
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def extract_section_text(full_text: str, start_marker: str, end_marker: str) -> str:
    """
    Return the slice of full_text between start_marker and end_marker.
    Returns empty string if not found.
    """
    start = full_text.find(start_marker)
    if start == -1:
        return ""
    end = full_text.find(end_marker, start)
    if end == -1:
        return full_text[start:]
    return full_text[start:end]


# ==============================================================
# PART-SPECIFIC PARSERS
# ==============================================================

# ---- PART A / Part A1 / Part A2  (TDS / Salary / Other than Salary) ------

def parse_part_a_transactions(section_text: str) -> pd.DataFrame:
    """
    Parse TDS transaction rows from Part A (and similar Part A1 / Part A2) text.
    Returns a DataFrame.
    """
    transactions = []
    current_name = ""
    current_tan = ""

    company_pattern = re.compile(
        r'([A-Z0-9 &.,()\'/-]+?)\s+([A-Z]{4}\d{5}[A-Z])\b'
    )
    transaction_pattern = re.compile(
        r'(\d+)\s+'
        r'([0-9]{1,4}[A-Z0-9]{0,5})\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'([A-Z])\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'(-|[A-Z0-9]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )

    for line in section_text.split("\n"):
        cm = company_pattern.search(line)
        if cm:
            current_name = cm.group(1).strip()
            current_tan  = cm.group(2)

        tm = transaction_pattern.search(line)
        if tm:
            transactions.append([
                tm.group(1),
                current_name,
                current_tan,
                tm.group(2),
                tm.group(3),
                tm.group(4),
                tm.group(5),
                tm.group(6),
                safe_float(tm.group(7)),
                safe_float(tm.group(8)),
                safe_float(tm.group(9)),
            ])

    return pd.DataFrame(transactions, columns=[
        "S.No", "Deductor Name", "TAN", "Section",
        "Transaction Date", "Booking Status", "Booking Date", "Remarks",
        "Amount Paid/Credited", "Tax Deducted", "TDS Deposited",
    ])


def parse_part_a_summary(section_text: str, section_lookup: dict) -> pd.DataFrame:
    """
    Parse the TDS Summary table from Part A (or similar section).
    section_lookup maps TAN → section code from transaction rows.
    """
    summary = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z0-9 &.,()\'/-]+?)\s+'
        r'([A-Z]{4}\d{5}[A-Z])\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        summary.append([
            m.group(1),
            section_lookup.get(m.group(3), ""),
            m.group(2).strip(),
            m.group(3),
            safe_float(m.group(4)),
            safe_float(m.group(5)),
            safe_float(m.group(6)),
        ])
    return pd.DataFrame(summary, columns=[
        "S.No", "Section", "Name", "TAN",
        "Amount Paid/Credited", "TDS Deducted", "TDS Deposited",
    ])


def build_section_lookup(transactions_df: pd.DataFrame) -> dict:
    """Build TAN → Section mapping from a transactions DataFrame."""
    lookup = {}
    if transactions_df.empty:
        return lookup
    for _, row in transactions_df.iterrows():
        tan = row.get("TAN", "")
        sec = row.get("Section", "")
        if tan and tan not in lookup:
            lookup[tan] = sec
    return lookup


# ---- PART B  (TCS – Tax Collected at Source) ----------------------------

def parse_part_b_transactions(section_text: str) -> pd.DataFrame:
    """
    Part B has TCS transactions.  Column structure is similar to Part A
    but uses "Tax Collected" instead of "Tax Deducted".
    """
    transactions = []
    current_name = ""
    current_tan  = ""

    company_pattern = re.compile(
        r'([A-Z0-9 &.,()\'/-]+?)\s+([A-Z]{4}\d{5}[A-Z])\b'
    )
    # TCS rows typically have the same shape as TDS rows
    trx_pattern = re.compile(
        r'(\d+)\s+'
        r'([0-9]{1,4}[A-Z0-9]{0,5})\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'([A-Z])\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'(-|[A-Z0-9]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )

    for line in section_text.split("\n"):
        cm = company_pattern.search(line)
        if cm:
            current_name = cm.group(1).strip()
            current_tan  = cm.group(2)

        tm = trx_pattern.search(line)
        if tm:
            transactions.append([
                tm.group(1),
                current_name,
                current_tan,
                tm.group(2),
                tm.group(3),
                tm.group(4),
                tm.group(5),
                tm.group(6),
                safe_float(tm.group(7)),
                safe_float(tm.group(8)),
                safe_float(tm.group(9)),
            ])

    return pd.DataFrame(transactions, columns=[
        "S.No", "Collector Name", "TAN", "Section",
        "Transaction Date", "Booking Status", "Booking Date", "Remarks",
        "Amount Paid/Debited", "Tax Collected", "TCS Deposited",
    ])


def parse_part_b_summary(section_text: str, section_lookup: dict) -> pd.DataFrame:
    summary = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z0-9 &.,()\'/-]+?)\s+'
        r'([A-Z]{4}\d{5}[A-Z])\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        summary.append([
            m.group(1),
            section_lookup.get(m.group(3), ""),
            m.group(2).strip(),
            m.group(3),
            safe_float(m.group(4)),
            safe_float(m.group(5)),
            safe_float(m.group(6)),
        ])
    return pd.DataFrame(summary, columns=[
        "S.No", "Section", "Collector Name", "TAN",
        "Amount Paid/Debited", "Tax Collected", "TCS Deposited",
    ])


# ---- PART C  (Tax Paid by Taxpayer – Advance Tax / Self Assessment) ------

def parse_part_c_transactions(section_text: str) -> pd.DataFrame:
    """
    Part C rows: BSR Code | Date of Deposit | Challan No | Section | Amount | Remarks
    """
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'                          # S.No
        r'(\d{7})\s+'                        # BSR Code (7 digits)
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'     # Date of Deposit
        r'(\d+)\s+'                          # Challan No
        r'([0-9]{3}[A-Z0-9]{0,3})\s+'       # Section / Minor Head
        r'([\d,.]+)\s+'                      # Amount
        r'(-|[A-Z0-9 ]+)'                    # Remarks
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3),
            m.group(4),
            m.group(5),
            safe_float(m.group(6)),
            m.group(7).strip(),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "BSR Code", "Date of Deposit", "Challan No",
        "Section / Minor Head", "Amount (Rs.)", "Remarks",
    ])


def parse_part_c_summary(section_text: str) -> pd.DataFrame:
    """
    Part C summary – total by minor head.
    """
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Za-z0-9 &.,()/-]+?)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2).strip(),
            safe_float(m.group(3)),
        ])
    return pd.DataFrame(rows, columns=["S.No", "Minor Head", "Amount (Rs.)"])


# ---- PART D  (Refund) ----------------------------------------------------

def parse_part_d_transactions(section_text: str) -> pd.DataFrame:
    """
    Part D: Refund information.
    Columns: S.No | Assessment Year | Mode | Amount | Interest | Remarks
    """
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'(\d{4}-\d{2,4})\s+'               # AY  e.g. 2023-24
        r'(NECS|CHEQUE|ECS|RTGS|[A-Z]+)\s+' # Mode
        r'([\d,.]+)\s+'                      # Refund Amount
        r'([\d,.]+)\s*'                      # Interest
        r'(-|[A-Z0-9 ]+)?'                   # Remarks
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3),
            safe_float(m.group(4)),
            safe_float(m.group(5)),
            (m.group(6) or "").strip(),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "Assessment Year", "Mode",
        "Refund Amount (Rs.)", "Interest (Rs.)", "Remarks",
    ])


def parse_part_d_summary(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'(\d{4}-\d{2,4})\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            safe_float(m.group(3)),
            safe_float(m.group(4)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "Assessment Year", "Refund Amount (Rs.)", "Interest (Rs.)",
    ])


# ---- PART E  (AIR / SFT – Statement of Financial Transactions) -----------

def parse_part_e_transactions(section_text: str) -> pd.DataFrame:
    """
    Part E / SFT rows:
    S.No | SFT Code | Type of Transaction | Amount | Remarks
    """
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'(SFT-\d{3}|[A-Z]{2,4}\d*)\s+'    # SFT Code
        r'([A-Za-z0-9 &.,()/-]+?)\s+'       # Type of transaction
        r'([\d,.]+)\s*'                      # Amount
        r'(-|[A-Z0-9 ]+)?'                   # Remarks
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            safe_float(m.group(4)),
            (m.group(5) or "").strip(),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "SFT Code", "Type of Transaction",
        "Amount (Rs.)", "Remarks",
    ])


def parse_part_e_summary(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'(SFT-\d{3}|[A-Z]{2,4}\d*)\s+'
        r'([A-Za-z0-9 &.,()/-]+?)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            safe_float(m.group(4)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "SFT Code", "Type of Transaction", "Amount (Rs.)",
    ])


# ---- PART F  (TDS on Sale of Immovable Property – 26QB / 26QC / 26QD) ---

def parse_part_f_transactions(section_text: str) -> pd.DataFrame:
    """
    Part F rows: S.No | PAN of Buyer | Name of Buyer | Acknowledgement No |
                 Transaction Date | Booking Date | Amount | Tax Deducted | TDS Deposited
    """
    rows = []
    # PAN of Buyer is like any PAN: AAAAA9999A
    pan_pattern = re.compile(r'[A-Z]{5}\d{4}[A-Z]')

    trx_pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{5}\d{4}[A-Z])\s+'          # Buyer PAN
        r'([A-Z0-9 &.,()/-]+?)\s+'           # Buyer Name
        r'(\d{10,15})\s+'                    # Acknowledgement No
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'     # Transaction Date
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'     # Booking Date
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in trx_pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            m.group(4),
            m.group(5),
            m.group(6),
            safe_float(m.group(7)),
            safe_float(m.group(8)),
            safe_float(m.group(9)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "PAN of Buyer", "Name of Buyer", "Acknowledgement No",
        "Transaction Date", "Booking Date",
        "Amount Paid/Credited", "Tax Deducted", "TDS Deposited",
    ])


def parse_part_f_summary(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{5}\d{4}[A-Z])\s+'
        r'([A-Z0-9 &.,()/-]+?)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            safe_float(m.group(4)),
            safe_float(m.group(5)),
            safe_float(m.group(6)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "PAN of Buyer", "Name of Buyer",
        "Amount Paid/Credited", "Tax Deducted", "TDS Deposited",
    ])


# ---- PART G  (TDS Defaults) -----------------------------------------------

def parse_part_g(section_text: str) -> pd.DataFrame:
    """
    Part G: TDS defaults / short-payment / short-deduction notices.
    """
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{4}\d{5}[A-Z])\s+'          # TAN
        r'([A-Za-z0-9 &.,()/-]+?)\s+'        # Deductor
        r'([A-Z0-9]+)\s+'                    # Default Type / Notice
        r'([\d,.]+)\s*'                      # Demand Amount
        r'(-|[A-Z0-9 ]+)?'                   # Remarks
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            m.group(4),
            safe_float(m.group(5)),
            (m.group(6) or "").strip(),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "TAN", "Deductor Name", "Default Type",
        "Demand Amount (Rs.)", "Remarks",
    ])


# ---- PART H  (Turnover Reported in GSTR) ----------------------------------

def parse_part_h(section_text: str) -> pd.DataFrame:
    """
    Part H: GST Turnover details.
    Columns: S.No | GSTIN | Name | FY | Turnover (Rs.)
    """
    rows = []
    gstin_pat = re.compile(r'\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]')
    pattern = re.compile(
        r'(\d+)\s+'
        r'(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9])\s+'   # GSTIN 15 chars
        r'([A-Z0-9 &.,()/-]+?)\s+'
        r'(\d{4}-\d{2,4})\s+'                            # FY
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1),
            m.group(2),
            m.group(3).strip(),
            m.group(4),
            safe_float(m.group(5)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "GSTIN", "Name", "Financial Year", "Turnover (Rs.)",
    ])


# ---- PART I  (TDS on Rent of Property – 26QC) ----------------------------

def parse_part_i_transactions(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{5}\d{4}[A-Z])\s+'           # Tenant PAN
        r'([A-Za-z0-9 &.,()/-]+?)\s+'         # Tenant Name
        r'(\d{10,15})\s+'                     # Challan / Ack No
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'      # Transaction Date
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'      # Booking Date
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1), m.group(2), m.group(3).strip(), m.group(4),
            m.group(5), m.group(6),
            safe_float(m.group(7)),
            safe_float(m.group(8)),
            safe_float(m.group(9)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "PAN of Tenant", "Name of Tenant", "Acknowledgement No",
        "Transaction Date", "Booking Date",
        "Rent Paid", "Tax Deducted", "TDS Deposited",
    ])


def parse_part_i_summary(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{5}\d{4}[A-Z])\s+'
        r'([A-Za-z0-9 &.,()/-]+?)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1), m.group(2), m.group(3).strip(),
            safe_float(m.group(4)),
            safe_float(m.group(5)),
            safe_float(m.group(6)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "PAN of Tenant", "Name of Tenant",
        "Rent Paid", "Tax Deducted", "TDS Deposited",
    ])


# ---- PART J  (TDS on Payment to Senior Citizen – 194P) -------------------

def parse_part_j_transactions(section_text: str) -> pd.DataFrame:
    rows = []
    pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z]{4}\d{5}[A-Z])\s+'           # Bank TAN
        r'([A-Za-z0-9 &.,()/-]+?)\s+'         # Bank Name
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'      # Date
        r'([A-Z])\s+'                         # Status
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'      # Booking Date
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in pattern.finditer(section_text):
        rows.append([
            m.group(1), m.group(2), m.group(3).strip(),
            m.group(4), m.group(5), m.group(6),
            safe_float(m.group(7)),
            safe_float(m.group(8)),
            safe_float(m.group(9)),
        ])
    return pd.DataFrame(rows, columns=[
        "S.No", "TAN of Bank", "Bank Name",
        "Transaction Date", "Booking Status", "Booking Date",
        "Amount Paid", "Tax Deducted", "TDS Deposited",
    ])


# ==============================================================
# MAIN APPLICATION
# ==============================================================

if uploaded_file:

    with st.spinner("Reading PDF…"):
        with pdfplumber.open(uploaded_file) as pdf:
            full_text = ""
            all_tables = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
                tables = page.extract_tables()
                if tables:
                    all_tables.extend(tables)

    # ------------------------------------------------------------------
    # Assessee Details
    # ------------------------------------------------------------------

    assessee = {}

    pan_match = re.search(
        r"Permanent Account Number \(PAN\)\s+([A-Z0-9]+)", full_text
    )
    name_match = re.search(
        r"Name of Assessee\s+(.+?)\n", full_text
    )
    fy_match = re.search(
        r"Financial Year\s+([0-9\-]+)", full_text
    )
    ay_match = re.search(
        r"Assessment Year\s+([0-9\-]+)", full_text
    )
    status_match = re.search(
        r"Current Status of PAN\s+(.+?)\s+Financial Year", full_text
    )
    address_match = re.search(
        r"Address of Assessee\s+(.+?)Above data", full_text, re.S
    )

    assessee["PAN"] = pan_match.group(1) if pan_match else ""
    assessee["Name"] = name_match.group(1).strip() if name_match else ""
    assessee["Financial Year"] = fy_match.group(1) if fy_match else ""
    assessee["Assessment Year"] = ay_match.group(1) if ay_match else ""
    assessee["PAN Status"] = status_match.group(1).strip() if status_match else ""
    if address_match:
        assessee["Address"] = address_match.group(1).replace("\n", " ").strip()
    else:
        assessee["Address"] = ""

    assessee_df = pd.DataFrame(list(assessee.items()), columns=["Field", "Value"])

    # ------------------------------------------------------------------
    # Locate section boundaries in the PDF text
    # ------------------------------------------------------------------

    PART_MARKERS = {
        "A":   ("PART A - DETAILS OF TAX DEDUCTED AT SOURCE",
                 "PART A1"),
        "A1":  ("PART A1 - DETAILS OF TAX DEDUCTED AT SOURCE FOR 15G/15H",
                 "PART A2"),
        "A2":  ("PART A2 - DETAILS OF TAX DEDUCTED AT SOURCE ON SALE OF IMMOVABLE PROPERTY",
                 "PART B"),
        "B":   ("PART B - DETAILS OF TAX COLLECTED AT SOURCE",
                 "PART C"),
        "C":   ("PART C - DETAILS OF TAX PAID",
                 "PART D"),
        "D":   ("PART D - DETAILS OF PAID REFUND",
                 "PART E"),
        "E":   ("PART E - DETAILS OF AIR TRANSACTION",
                 "PART F"),
        "F":   ("PART F - DETAILS OF TAX DEDUCTED AT SOURCE ON SALE OF IMMOVABLE PROPERTY",
                 "PART G"),
        "G":   ("PART G - TDS DEFAULTS",
                 "PART H"),
        "H":   ("PART H - DETAILS OF TURNOVER AS PER GSTR-3B",
                 "PART I"),
        "I":   ("PART I - DETAILS OF TAX DEDUCTED AT SOURCE ON SALE / RENT OF IMMOVABLE PROPERTY",
                 "PART J"),
        "J":   ("PART J - DETAILS OF TAX DEDUCTED AT SOURCE UNDER SECTION 194P",
                 "\x00"),          # last section – go to end
    }

    def get_part_text(part_key: str) -> str:
        if part_key not in PART_MARKERS:
            return ""
        start_marker, end_marker = PART_MARKERS[part_key]
        return extract_section_text(full_text, start_marker, end_marker)

    # ----------------------------------------------------------------
    # PART A  – TDS (original logic, preserved exactly)
    # ----------------------------------------------------------------

    part_a_text = get_part_text("A") or full_text   # fall back to full text

    transactions = []
    current_name = ""
    current_tan  = ""
    lines = part_a_text.split("\n")

    company_pattern = re.compile(
        r'([A-Z0-9 &.,()/-]+?)\s+([A-Z]{4}\d{5}[A-Z])'
    )
    transaction_pattern = re.compile(
        r'(\d+)\s+'
        r'([0-9]{3}[A-Z0-9]{0,3})\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'([A-Z])\s+'
        r'(\d{2}-[A-Za-z]{3}-\d{4})\s+'
        r'-\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )

    for line in lines:
        company_match = company_pattern.search(line)
        if company_match:
            current_name = company_match.group(1).strip()
            current_tan  = company_match.group(2)
        trx_match = transaction_pattern.search(line)
        if trx_match:
            transactions.append([
                trx_match.group(1),
                current_name,
                current_tan,
                trx_match.group(2),
                trx_match.group(3),
                trx_match.group(4),
                trx_match.group(5),
                "-",
                float(trx_match.group(6).replace(",", "")),
                float(trx_match.group(7).replace(",", "")),
                float(trx_match.group(8).replace(",", "")),
            ])

    transaction_df = pd.DataFrame(transactions, columns=[
        "S.No", "Deductor Name", "TAN", "Section",
        "Transaction Date", "Booking Status", "Booking Date", "Remarks",
        "Amount Paid/Credited", "Tax Deducted", "TDS Deposited",
    ])

    # Section lookup for summary
    section_lookup = {}
    for row in transactions:
        tan = row[2]
        sec = row[3]
        if tan not in section_lookup:
            section_lookup[tan] = sec

    # TDS Summary (original logic, preserved exactly)
    tds_summary = []
    summary_pattern = re.compile(
        r'(\d+)\s+'
        r'([A-Z0-9 &.,()/-]+?)\s+'
        r'([A-Z]{4}\d{5}[A-Z])\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)\s+'
        r'([\d,.]+)'
    )
    for m in summary_pattern.findall(part_a_text):
        sno        = m[0]
        name       = m[1].strip()
        tan        = m[2]
        amount     = float(m[3].replace(",", ""))
        tds        = float(m[4].replace(",", ""))
        deposited  = float(m[5].replace(",", ""))
        tds_summary.append([sno, section_lookup.get(tan, ""), name, tan,
                             amount, tds, deposited])

    tds_df = pd.DataFrame(tds_summary, columns=[
        "S.No", "Section", "Name", "TAN",
        "Amount Paid/Credited", "TDS Deducted", "TDS Deposited",
    ])

    # ----------------------------------------------------------------
    # PART A1  – TDS 15G / 15H
    # ----------------------------------------------------------------
    part_a1_text = get_part_text("A1")
    part_a1_trx_df  = pd.DataFrame()
    part_a1_sum_df  = pd.DataFrame()
    if part_a1_text:
        part_a1_trx_df = parse_part_a_transactions(part_a1_text)
        sl_a1 = build_section_lookup(part_a1_trx_df)
        part_a1_sum_df = parse_part_a_summary(part_a1_text, sl_a1)

    # ----------------------------------------------------------------
    # PART A2  – TDS on Immovable Property (buyer-side)
    # ----------------------------------------------------------------
    part_a2_text = get_part_text("A2")
    part_a2_trx_df  = pd.DataFrame()
    part_a2_sum_df  = pd.DataFrame()
    if part_a2_text:
        part_a2_trx_df = parse_part_a_transactions(part_a2_text)
        sl_a2 = build_section_lookup(part_a2_trx_df)
        part_a2_sum_df = parse_part_a_summary(part_a2_text, sl_a2)

    # ----------------------------------------------------------------
    # PART B  – TCS
    # ----------------------------------------------------------------
    part_b_text = get_part_text("B")
    part_b_trx_df  = pd.DataFrame()
    part_b_sum_df  = pd.DataFrame()
    if part_b_text:
        part_b_trx_df = parse_part_b_transactions(part_b_text)
        sl_b = build_section_lookup(part_b_trx_df)
        part_b_sum_df = parse_part_b_summary(part_b_text, sl_b)

    # ----------------------------------------------------------------
    # PART C  – Tax Paid (Advance / Self-Assessment)
    # ----------------------------------------------------------------
    part_c_text = get_part_text("C")
    part_c_trx_df  = pd.DataFrame()
    part_c_sum_df  = pd.DataFrame()
    if part_c_text:
        part_c_trx_df = parse_part_c_transactions(part_c_text)
        part_c_sum_df = parse_part_c_summary(part_c_text)

    # ----------------------------------------------------------------
    # PART D  – Refund
    # ----------------------------------------------------------------
    part_d_text = get_part_text("D")
    part_d_trx_df  = pd.DataFrame()
    part_d_sum_df  = pd.DataFrame()
    if part_d_text:
        part_d_trx_df = parse_part_d_transactions(part_d_text)
        part_d_sum_df = parse_part_d_summary(part_d_text)

    # ----------------------------------------------------------------
    # PART E  – AIR / SFT
    # ----------------------------------------------------------------
    part_e_text = get_part_text("E")
    part_e_trx_df  = pd.DataFrame()
    part_e_sum_df  = pd.DataFrame()
    if part_e_text:
        part_e_trx_df = parse_part_e_transactions(part_e_text)
        part_e_sum_df = parse_part_e_summary(part_e_text)

    # ----------------------------------------------------------------
    # PART F  – TDS on Sale of Immovable Property (seller-side 26QB)
    # ----------------------------------------------------------------
    part_f_text = get_part_text("F")
    part_f_trx_df  = pd.DataFrame()
    part_f_sum_df  = pd.DataFrame()
    if part_f_text:
        part_f_trx_df = parse_part_f_transactions(part_f_text)
        part_f_sum_df = parse_part_f_summary(part_f_text)

    # ----------------------------------------------------------------
    # PART G  – TDS Defaults
    # ----------------------------------------------------------------
    part_g_text = get_part_text("G")
    part_g_df = pd.DataFrame()
    if part_g_text:
        part_g_df = parse_part_g(part_g_text)

    # ----------------------------------------------------------------
    # PART H  – GST Turnover
    # ----------------------------------------------------------------
    part_h_text = get_part_text("H")
    part_h_df = pd.DataFrame()
    if part_h_text:
        part_h_df = parse_part_h(part_h_text)

    # ----------------------------------------------------------------
    # PART I  – TDS on Rent (26QC)
    # ----------------------------------------------------------------
    part_i_text = get_part_text("I")
    part_i_trx_df  = pd.DataFrame()
    part_i_sum_df  = pd.DataFrame()
    if part_i_text:
        part_i_trx_df = parse_part_i_transactions(part_i_text)
        part_i_sum_df = parse_part_i_summary(part_i_text)

    # ----------------------------------------------------------------
    # PART J  – TDS 194P (Senior Citizen)
    # ----------------------------------------------------------------
    part_j_text = get_part_text("J")
    part_j_trx_df = pd.DataFrame()
    if part_j_text:
        part_j_trx_df = parse_part_j_transactions(part_j_text)

    # ==================================================================
    # STREAMLIT PREVIEW
    # ==================================================================

    def show_section(title: str, df: pd.DataFrame):
        """Display a titled section only if the DataFrame has rows."""
        if df is not None and not df.empty:
            st.subheader(title)
            st.dataframe(df, use_container_width=True)

    st.subheader("Assessee Details")
    st.dataframe(assessee_df, use_container_width=True)

    # Part A
    show_section("Part A – TDS Summary", tds_df)
    show_section("Part A – TDS Transactions", transaction_df)

    # Part A1
    show_section("Part A1 – TDS (15G/15H) Summary", part_a1_sum_df)
    show_section("Part A1 – TDS (15G/15H) Transactions", part_a1_trx_df)

    # Part A2
    show_section("Part A2 – TDS on Immovable Property Summary", part_a2_sum_df)
    show_section("Part A2 – TDS on Immovable Property Transactions", part_a2_trx_df)

    # Part B
    show_section("Part B – TCS Summary", part_b_sum_df)
    show_section("Part B – TCS Transactions", part_b_trx_df)

    # Part C
    show_section("Part C – Tax Paid Summary", part_c_sum_df)
    show_section("Part C – Tax Paid Transactions", part_c_trx_df)

    # Part D
    show_section("Part D – Refund Summary", part_d_sum_df)
    show_section("Part D – Refund Details", part_d_trx_df)

    # Part E
    show_section("Part E – AIR / SFT Summary", part_e_sum_df)
    show_section("Part E – AIR / SFT Transactions", part_e_trx_df)

    # Part F
    show_section("Part F – TDS on Property Sale Summary", part_f_sum_df)
    show_section("Part F – TDS on Property Sale Transactions", part_f_trx_df)

    # Part G
    show_section("Part G – TDS Defaults", part_g_df)

    # Part H
    show_section("Part H – GST Turnover", part_h_df)

    # Part I
    show_section("Part I – TDS on Rent Summary", part_i_sum_df)
    show_section("Part I – TDS on Rent Transactions", part_i_trx_df)

    # Part J
    show_section("Part J – TDS 194P (Senior Citizen) Transactions", part_j_trx_df)

    # ==================================================================
    # EXCEL EXPORT
    # ==================================================================

    output = BytesIO()

    # Reset table counter for each export
    _table_counter[0] = 0

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ---- Assessee Details ----------------------------------------
        assessee_df.to_excel(writer, sheet_name="Assessee Details", index=False)
        ws_ad = writer.sheets["Assessee Details"]
        apply_sheet_formatting(ws_ad, table_name=_next_table_name("AssesseeDetails"))

        # ---- Part A Summary (original) --------------------------------
        if not tds_df.empty:
            tds_df.to_excel(writer, sheet_name="Part A Summary", index=False)
            ws = writer.sheets["Part A Summary"]
            num_cols  = [col_idx + 1 for col_idx, c in enumerate(tds_df.columns)
                         if tds_df[c].dtype in [float, int] and "S.No" not in c]
            total_pairs = [(i, get_column_letter(i)) for i in num_cols]
            apply_sheet_formatting(
                ws,
                numeric_cols=num_cols,
                total_col_indices=total_pairs,
                table_name=_next_table_name("PartASummary"),
            )

        # ---- Part A Transactions (original) ---------------------------
        if not transaction_df.empty:
            transaction_df.to_excel(writer, sheet_name="Part A Transactions", index=False)
            ws = writer.sheets["Part A Transactions"]
            num_idx = [col_idx + 1 for col_idx, c in enumerate(transaction_df.columns)
                       if transaction_df[c].dtype in [float, int] and "S.No" not in c]
            apply_sheet_formatting(
                ws,
                numeric_cols=num_idx,
                table_name=_next_table_name("PartATransactions"),
            )

        # ---- Helper: write a summary sheet with totals ----------------
        def _write_summary_sheet(df, sheet_name, table_prefix):
            if df is None or df.empty:
                return
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            num_cols = [i + 1 for i, c in enumerate(df.columns)
                        if df[c].dtype in [float, int] and c != "S.No"]
            total_pairs = [(i, get_column_letter(i)) for i in num_cols]
            apply_sheet_formatting(
                ws,
                numeric_cols=num_cols,
                total_col_indices=total_pairs,
                table_name=_next_table_name(table_prefix),
            )

        def _write_trx_sheet(df, sheet_name, table_prefix):
            if df is None or df.empty:
                return
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            num_cols = [i + 1 for i, c in enumerate(df.columns)
                        if df[c].dtype in [float, int] and c != "S.No"]
            apply_sheet_formatting(
                ws,
                numeric_cols=num_cols,
                table_name=_next_table_name(table_prefix),
            )

        # ---- Part A1 --------------------------------------------------
        _write_summary_sheet(part_a1_sum_df, "Part A1 Summary", "PartA1Summary")
        _write_trx_sheet(part_a1_trx_df, "Part A1 Transactions", "PartA1Trx")

        # ---- Part A2 --------------------------------------------------
        _write_summary_sheet(part_a2_sum_df, "Part A2 Summary", "PartA2Summary")
        _write_trx_sheet(part_a2_trx_df, "Part A2 Transactions", "PartA2Trx")

        # ---- Part B ---------------------------------------------------
        _write_summary_sheet(part_b_sum_df, "Part B Summary", "PartBSummary")
        _write_trx_sheet(part_b_trx_df, "Part B Transactions", "PartBTrx")

        # ---- Part C ---------------------------------------------------
        _write_summary_sheet(part_c_sum_df, "Part C Summary", "PartCSummary")
        _write_trx_sheet(part_c_trx_df, "Part C Transactions", "PartCTrx")

        # ---- Part D ---------------------------------------------------
        _write_summary_sheet(part_d_sum_df, "Part D Summary", "PartDSummary")
        _write_trx_sheet(part_d_trx_df, "Part D Details", "PartDTrx")

        # ---- Part E ---------------------------------------------------
        _write_summary_sheet(part_e_sum_df, "Part E Summary", "PartESummary")
        _write_trx_sheet(part_e_trx_df, "Part E Transactions", "PartETrx")

        # ---- Part F ---------------------------------------------------
        _write_summary_sheet(part_f_sum_df, "Part F Summary", "PartFSummary")
        _write_trx_sheet(part_f_trx_df, "Part F Transactions", "PartFTrx")

        # ---- Part G ---------------------------------------------------
        _write_trx_sheet(part_g_df, "Part G Defaults", "PartG")

        # ---- Part H ---------------------------------------------------
        _write_summary_sheet(part_h_df, "Part H GST Turnover", "PartH")

        # ---- Part I ---------------------------------------------------
        _write_summary_sheet(part_i_sum_df, "Part I Summary", "PartISummary")
        _write_trx_sheet(part_i_trx_df, "Part I Transactions", "PartITrx")

        # ---- Part J ---------------------------------------------------
        _write_trx_sheet(part_j_trx_df, "Part J Transactions", "PartJTrx")

        # ---- Preserve original TDS Summary Excel Table (back-compat) --
        # The original code added a Table named "TDSSummary" to the
        # "TDS Summary" sheet.  We replicate that on "Part A Summary".
        # (Already handled above via apply_sheet_formatting.)

    output.seek(0)

    st.download_button(
        "⬇️  Download Excel",
        output,
        file_name="26AS_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
